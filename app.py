import pandas as pd
import streamlit as st
from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment, PatternFill
import numpy as np
import time
import plotly.express as px
import plotly.graph_objects as go

st.set_page_config(page_title="PTD vs SDS Comparison Tool", layout="wide")


# ─────────────────────────────────────────────
# HELPER UTILITIES
# ─────────────────────────────────────────────

def find_column(df, target_name):
    target_lower = target_name.strip().lower()
    return next(
        (col for col in df.columns if col.strip().lower() == target_lower), None
    )


def reset_file(uploaded_file):
    try:
        uploaded_file.seek(0)
    except Exception:
        pass


def make_composite_key(item_name, form_label, item_group_label):
    item_name_s  = str(item_name).strip()        if pd.notna(item_name)        else ''
    form_label_s = str(form_label).strip()       if pd.notna(form_label)       else ''
    item_group_s = str(item_group_label).strip() if pd.notna(item_group_label) else ''
    if form_label_s:
        return (item_name_s, 'form', form_label_s)
    elif item_group_s:
        return (item_name_s, 'group', item_group_s)
    else:
        return (item_name_s, '', '')


def make_tab_label(item_name, form_label, item_group_label):
    item_name_s  = str(item_name).strip()        if pd.notna(item_name)        else ''
    form_label_s = str(form_label).strip()       if pd.notna(form_label)       else ''
    item_group_s = str(item_group_label).strip() if pd.notna(item_group_label) else ''
    if form_label_s:
        return f"{item_name_s} ({form_label_s})"
    elif item_group_s:
        return f"{item_name_s} ({item_group_s})"
    else:
        return item_name_s


# ─────────────────────────────────────────────
# FILE PARSING
# ─────────────────────────────────────────────

def parse_uploaded_file(uploaded_file, sheet_name='Form Definitions', is_ptd=False):
    try:
        reset_file(uploaded_file)
        if is_ptd and sheet_name == 'Form Definitions':
            df = pd.read_excel(
                uploaded_file, sheet_name=sheet_name,
                engine='openpyxl', header=1
            )
        else:
            df = pd.read_excel(
                uploaded_file, sheet_name=sheet_name, engine='openpyxl'
            )
        return df
    except ValueError:
        reset_file(uploaded_file)
        try:
            available = pd.ExcelFile(uploaded_file).sheet_names
        except Exception:
            available = []
        st.error(
            f"Error: Sheet '**{sheet_name}**' not found. "
            f"Available sheets: {available}"
        )
        return None
    except Exception as e:
        st.error(f"Error reading file: {e}")
        return None


# ─────────────────────────────────────────────
# DATA PROCESSING
# ─────────────────────────────────────────────

def convert_decimal_column(df):
    if df is None:
        return df
    decimal_column = find_column(df, 'decimal')
    if not decimal_column:
        return df

    def format_decimal(value):
        if pd.isna(value) or value == '' or value is None:
            return value
        try:
            str_value  = str(value).strip()
            if not str_value:
                return value
            float_value = float(str_value)
            if float_value == int(float_value):
                return f"{int(float_value)}.0"
            else:
                return str(float_value)
        except (ValueError, TypeError):
            return value

    df[decimal_column] = df[decimal_column].apply(format_decimal)
    return df


def process_choice_code_sheet(df, sheet_type="Codelist"):
    if df is None:
        return None, 0, 0
    choice_code_col = find_column(df, 'choice code')
    if choice_code_col:
        original_count = len(df)
        df = df[df[choice_code_col].notna()].copy()
        df = df[df[choice_code_col].astype(str).str.strip() != ''].copy()
        return df, original_count, len(df)
    return df, len(df), len(df)


def process_ptd_dataframe(df):
    if df is None:
        return None, 0, 0
    original_count = len(df)
    trial_column_names = [
        'Used in trial (Y, N, Mod)',
        'Used in trial (Y, N, Mod) ',
        ' Used in trial (Y, N, Mod)',
        ' Used in trial (Y, N, Mod) ',
        'Used in trial',
    ]
    trial_column = next(
        (col for col in trial_column_names if col in df.columns), None
    )
    if trial_column:
        normalized = df[trial_column].astype(str).str.strip().str.upper()
        df = df[normalized.isin(['Y', 'YES', 'MOD'])].copy()
        filtered_count = len(df)
    else:
        st.warning("⚠️ Column 'Used in trial (Y, N, Mod)' not found. Skipping filter.")
        filtered_count = original_count

    columns_to_remove_patterns = {
        'Modification comments + Highlight Cells where change made',
        'Library source',
        'Used in trial (Y, N, Mod)'
    }
    columns_to_remove = [
        col for col in df.columns
        if col.strip() in columns_to_remove_patterns
    ]
    df = df.drop(columns=columns_to_remove, errors='ignore')
    df = convert_decimal_column(df)
    return df, original_count, filtered_count


def process_sds_dataframe(df):
    if df is None:
        return df
    return convert_decimal_column(df)


# ─────────────────────────────────────────────
# LOOKUP / MATCHING
# ─────────────────────────────────────────────

def build_lookup_dictionaries(df):
    item_name_counts = {}
    for _, row in df.iterrows():
        item_name = row.get('Item Name')
        if pd.notna(item_name):
            key = str(item_name).strip()
            item_name_counts[key] = item_name_counts.get(key, 0) + 1

    duplicate_item_names = {k for k, v in item_name_counts.items() if v > 1}

    simple_dict    = {}
    composite_dict = {}

    for _, row in df.iterrows():
        item_name = row.get('Item Name')
        if pd.isna(item_name):
            continue
        item_name_s = str(item_name).strip()
        if item_name_s not in duplicate_item_names:
            simple_dict[item_name_s] = row
        else:
            form_label       = row.get('Form Label', '')
            item_group_label = row.get('Item Group Label', '')
            comp_key = make_composite_key(item_name_s, form_label, item_group_label)
            composite_dict[comp_key] = row

    return simple_dict, composite_dict, duplicate_item_names


def find_matching_rows_optimized(
    item_name,
    source_simple, source_composite, source_duplicates,
    target_simple, target_composite, target_duplicates,
    form_label=None, item_group_label=None
):
    item_name_s  = str(item_name).strip()
    is_duplicate = (
        item_name_s in source_duplicates or
        item_name_s in target_duplicates
    )
    if not is_duplicate:
        source_row = source_simple.get(item_name_s)
        target_row = target_simple.get(item_name_s)
        match_type = 'Item Name'
    else:
        comp_key   = make_composite_key(item_name_s, form_label, item_group_label)
        source_row = source_composite.get(comp_key)
        target_row = target_composite.get(comp_key)
        match_type = 'Item Name + Form Label'

    if source_row is not None or target_row is not None:
        return source_row, target_row, match_type
    return None, None, None


def get_unique_items(df, column_name='Item Name'):
    return set(df[column_name].dropna().unique())


def get_all_item_keys(df):
    seen_keys   = set()
    items       = []
    name_counts = {}

    for _, row in df.iterrows():
        n = row.get('Item Name')
        if pd.notna(n):
            s = str(n).strip()
            name_counts[s] = name_counts.get(s, 0) + 1

    duplicate_names = {k for k, v in name_counts.items() if v > 1}

    for _, row in df.iterrows():
        item_name = row.get('Item Name')
        if pd.isna(item_name):
            continue
        item_name_s      = str(item_name).strip()
        form_label       = row.get('Form Label', '')
        item_group_label = row.get('Item Group Label', '')

        if item_name_s in duplicate_names:
            comp_key  = make_composite_key(item_name_s, form_label, item_group_label)
            tab_label = make_tab_label(item_name_s, form_label, item_group_label)
        else:
            comp_key  = (item_name_s, '', '')
            tab_label = item_name_s

        if comp_key not in seen_keys:
            seen_keys.add(comp_key)
            items.append({
                'item_name':        item_name_s,
                'form_label':       str(form_label).strip()       if pd.notna(form_label)       else '',
                'item_group_label': str(item_group_label).strip() if pd.notna(item_group_label) else '',
                'tab_label':        tab_label,
                'comp_key':         comp_key,
            })
    return items


# ─────────────────────────────────────────────
# COMPARISON LOGIC
# ─────────────────────────────────────────────

def compare_values(val1, val2):
    try:
        val1_nan = pd.isna(val1)
    except (TypeError, ValueError):
        val1_nan = False
    try:
        val2_nan = pd.isna(val2)
    except (TypeError, ValueError):
        val2_nan = False

    if val1_nan and val2_nan:
        return 'match', '✓', ''
    if val1_nan and not val2_nan:
        return 'missing_source', '⚠', 'Missing in Source'
    if not val1_nan and val2_nan:
        return 'missing_target', '⚠', 'Missing in Target'
    if str(val1).strip() == str(val2).strip():
        return 'match', '✓', ''
    return 'mismatch', '✗', 'Values differ'


def compare_codelists(
    codelist_name, source_codelists_df, target_codelists_df,
    source_name, target_name, list_type="Codelist"
):
    if source_codelists_df is None or target_codelists_df is None:
        return None
    if pd.isna(codelist_name) or str(codelist_name).strip() == '':
        return None

    source_name_col = find_column(source_codelists_df, 'name')
    target_name_col = find_column(target_codelists_df, 'name')
    if not source_name_col or not target_name_col:
        return None

    cl_name_stripped = str(codelist_name).strip()
    source_codelist  = source_codelists_df[
        source_codelists_df[source_name_col].astype(str).str.strip() == cl_name_stripped
    ].copy()
    target_codelist  = target_codelists_df[
        target_codelists_df[target_name_col].astype(str).str.strip() == cl_name_stripped
    ].copy()

    if source_codelist.empty and target_codelist.empty:
        return {'status': 'not_found',
                'message': f"{list_type} '{codelist_name}' not found in either file",
                'matches': 0, 'mismatches': 0, 'details': [], 'type': list_type}
    if source_codelist.empty:
        return {'status': 'missing_source',
                'message': f"{list_type} '{codelist_name}' not found in {source_name}",
                'matches': 0, 'mismatches': len(target_codelist),
                'details': [], 'type': list_type}
    if target_codelist.empty:
        return {'status': 'missing_target',
                'message': f"{list_type} '{codelist_name}' not found in {target_name}",
                'matches': 0, 'mismatches': len(source_codelist),
                'details': [], 'type': list_type}

    source_cols = {}
    target_cols = {}
    for col in source_codelist.columns:
        cl = col.strip().lower()
        if cl in ('choice code', 'choice label'):
            source_cols[cl] = col
    for col in target_codelist.columns:
        cl = col.strip().lower()
        if cl in ('choice code', 'choice label'):
            target_cols[cl] = col

    matches = mismatches = 0
    details = []

    if 'choice code' in source_cols and 'choice code' in target_cols:
        source_codes = set(source_codelist[source_cols['choice code']].astype(str).str.strip())
        target_codes = set(target_codelist[target_cols['choice code']].astype(str).str.strip())

        for code in source_codes.union(target_codes):
            src_rows = source_codelist[
                source_codelist[source_cols['choice code']].astype(str).str.strip() == code]
            tgt_rows = target_codelist[
                target_codelist[target_cols['choice code']].astype(str).str.strip() == code]

            src_label = (src_rows.iloc[0][source_cols['choice label']]
                         if not src_rows.empty and 'choice label' in source_cols else '')
            tgt_label = (tgt_rows.iloc[0][target_cols['choice label']]
                         if not tgt_rows.empty and 'choice label' in target_cols else '')

            if src_rows.empty:
                mismatches += 1
                details.append({'name': codelist_name, 'choice_code': code,
                                 'source_label': 'Missing', 'target_label': tgt_label,
                                 'status': 'missing_source', 'type': list_type})
            elif tgt_rows.empty:
                mismatches += 1
                details.append({'name': codelist_name, 'choice_code': code,
                                 'source_label': src_label, 'target_label': 'Missing',
                                 'status': 'missing_target', 'type': list_type})
            elif str(src_label).strip() == str(tgt_label).strip():
                matches += 1
                details.append({'name': codelist_name, 'choice_code': code,
                                 'source_label': src_label, 'target_label': tgt_label,
                                 'status': 'match', 'type': list_type})
            else:
                mismatches += 1
                details.append({'name': codelist_name, 'choice_code': code,
                                 'source_label': src_label, 'target_label': tgt_label,
                                 'status': 'mismatch', 'type': list_type})

    status = 'match' if mismatches == 0 else 'mismatch'
    return {
        'status': status,
        'message': f"{list_type} '{codelist_name}': {matches} matches, {mismatches} mismatches",
        'matches': matches, 'mismatches': mismatches,
        'details': details, 'type': list_type
    }


def create_comparison_dataframe(
    source_row, target_row, source_columns,
    source_name, target_name,
    codelist_comparison=None, unit_codelist_comparison=None
):
    IGNORE_COLUMNS  = {'Definition Last Modified', 'Relationship Last Modified'}
    comparison_data = []

    source_values = (source_row if source_row is not None
                     else pd.Series([None] * len(source_columns), index=source_columns))
    target_values = (target_row if target_row is not None
                     else pd.Series([None] * len(source_columns), index=source_columns))

    for col in source_columns:
        if col in IGNORE_COLUMNS:
            continue
        source_value = source_values.get(col)
        target_value = target_values.get(col) if col in target_values.index else None

        status, symbol, note = compare_values(source_value, target_value)
        note = note.replace('Source', source_name).replace('Target', target_name)

        comparison_data.append({
            'Column Name':          col,
            f'{source_name} Value': source_value if not pd.isna(source_value) else '',
            f'{target_name} Value': (target_value if not pd.isna(target_value)
                                     else f'Column not in {target_name}'),
            'Status': symbol, 'Match': status, 'Note': note
        })

    for cl_comp, prefix in [
        (codelist_comparison,      'Codelist'),
        (unit_codelist_comparison, 'Unit Codelist'),
    ]:
        if cl_comp and cl_comp.get('details'):
            for detail in cl_comp['details']:
                status_map = {
                    'match':          ('match',          '✓', ''),
                    'mismatch':       ('mismatch',       '✗', f'{prefix} labels differ'),
                    'missing_source': ('missing_source', '⚠', f'Choice Code missing in {source_name}'),
                    'missing_target': ('missing_target', '⚠', f'Choice Code missing in {target_name}'),
                }
                status, symbol, note = status_map.get(
                    detail['status'], ('mismatch', '✗', 'Unknown'))
                comparison_data.append({
                    'Column Name':          f"{prefix}: {detail['name']} | Code: {detail['choice_code']}",
                    f'{source_name} Value': detail['source_label'],
                    f'{target_name} Value': detail['target_label'],
                    'Status': symbol, 'Match': status, 'Note': note
                })

    return pd.DataFrame(comparison_data)


def highlight_differences(row):
    match_type = row['Match']
    if match_type == 'match':
        return ['background-color: #90EE90'] * len(row)
    elif match_type == 'mismatch':
        return ['background-color: #FFB6C1'] * len(row)
    elif 'missing' in match_type:
        return ['background-color: #FFE4B5'] * len(row)
    return [''] * len(row)


# ─────────────────────────────────────────────
# ISSUE PARSER
# ─────────────────────────────────────────────

def parse_issue_fields(match_status, col_name, src_val, tgt_val,
                        source_name, target_name):
    col_lower = str(col_name).strip().lower()

    if match_status == 'item_not_found_source':
        return (f"Item Not Found in {source_name}", '',
                f"Item '{src_val}' exists in {target_name} but not in {source_name}")
    if match_status == 'item_not_found_target':
        return (f"Item Not Found in {target_name}", '',
                f"Item '{src_val}' exists in {source_name} but not in {target_name}")

    is_unit_codelist = col_lower.startswith('unit codelist:')
    is_codelist      = col_lower.startswith('codelist:')

    if is_codelist or is_unit_codelist:
        prefix       = 'Unit Codelist' if is_unit_codelist else 'Codelist'
        issue_column = col_name.split(':', 1)[1].strip() if ':' in col_name else col_name
        if match_status == 'mismatch':
            return (f"{prefix} Mismatch", issue_column,
                    f"{source_name}: '{src_val}' vs {target_name}: '{tgt_val}'")
        elif match_status == 'missing_source':
            return (f"{prefix} Missing in {source_name}", issue_column,
                    f"Choice Code present in {target_name} but missing in {source_name}")
        elif match_status == 'missing_target':
            return (f"{prefix} Missing in {target_name}", issue_column,
                    f"Choice Code present in {source_name} but missing in {target_name}")

    if match_status == 'mismatch':
        return ('Value Mismatch', col_name,
                f"{source_name}: '{src_val}' | {target_name}: '{tgt_val}'")
    elif match_status == 'missing_target':
        return (f"Missing in {target_name}", col_name,
                f"Present in {source_name} ('{src_val}') but absent in {target_name}")
    elif match_status == 'missing_source':
        return (f"Missing in {source_name}", col_name,
                f"Present in {target_name} ('{tgt_val}') but absent in {source_name}")

    return ('Other', col_name, f"{match_status}")


# ─────────────────────────────────────────────
# STATISTICS
# ─────────────────────────────────────────────

def compute_statistics(all_comparisons, missing_in_source, missing_in_target,
                        source_name, target_name):
    total_compared         = len(all_comparisons)
    total_100pct           = 0
    total_mismatches       = 0
    total_missing          = 0
    total_columns_compared = 0
    total_matched_columns  = 0
    issue_category_counts  = {}
    form_label_stats       = {}

    for tab_label, data in all_comparisons.items():
        comp_df   = data['comparison_df']
        mc        = comp_df['Match'].value_counts()
        total     = len(comp_df)
        matches_n = mc.get('match', 0)
        pct       = (matches_n / total * 100) if total > 0 else 0

        total_columns_compared += total
        total_matched_columns  += matches_n

        if pct == 100.0:
            total_100pct += 1
        if mc.get('mismatch', 0) > 0:
            total_mismatches += 1
        if mc.get('missing_source', 0) + mc.get('missing_target', 0) > 0:
            total_missing += 1

        issue_rows = comp_df[comp_df['Match'] != 'match']
        for _, row in issue_rows.iterrows():
            col_name     = row['Column Name']
            src_val      = row[f'{source_name} Value']
            tgt_val      = row[f'{target_name} Value']
            match_status = row['Match']
            cat, _, _    = parse_issue_fields(
                match_status, col_name, src_val, tgt_val, source_name, target_name
            )
            issue_category_counts[cat] = issue_category_counts.get(cat, 0) + 1

        form_label = data.get('form_label', '') or 'No Form Label'
        if form_label not in form_label_stats:
            form_label_stats[form_label] = {'total': 0, 'match_100': 0, 'with_issues': 0}
        form_label_stats[form_label]['total'] += 1
        if pct == 100.0:
            form_label_stats[form_label]['match_100'] += 1
        else:
            form_label_stats[form_label]['with_issues'] += 1

    for _ in missing_in_target:
        cat = f'Item Not Found in {target_name}'
        issue_category_counts[cat] = issue_category_counts.get(cat, 0) + 1
    for _ in missing_in_source:
        cat = f'Item Not Found in {source_name}'
        issue_category_counts[cat] = issue_category_counts.get(cat, 0) + 1

    return {
        'total_compared':         total_compared,
        'total_100pct':           total_100pct,
        'total_with_issues':      total_compared - total_100pct,
        'total_mismatches':       total_mismatches,
        'total_missing':          total_missing,
        'total_missing_items':    len(missing_in_source) + len(missing_in_target),
        'missing_in_source':      len(missing_in_source),
        'missing_in_target':      len(missing_in_target),
        'total_columns_compared': total_columns_compared,
        'total_matched_columns':  total_matched_columns,
        'overall_match_pct':      (total_matched_columns / total_columns_compared * 100
                                   if total_columns_compared > 0 else 0),
        'issue_category_counts':  issue_category_counts,
        'form_label_stats':       form_label_stats,
    }


def render_statistics(stats, source_name, target_name):
    st.markdown("---")
    st.subheader("📈 Comparison Statistics")

    st.markdown("##### 📊 Overview")
    m1, m2, m3, m4, m5, m6 = st.columns(6)
    m1.metric("Total Items Compared",      stats['total_compared'])
    m2.metric("✅ 100% Match",             stats['total_100pct'],
              delta=f"{(stats['total_100pct']/stats['total_compared']*100):.1f}%"
              if stats['total_compared'] > 0 else "0%")
    m3.metric("⚠️ Items with Issues",      stats['total_with_issues'],  delta_color="inverse")
    m4.metric("🔴 Items with Mismatches",  stats['total_mismatches'],   delta_color="inverse")
    m5.metric("🟡 Items with Missing",     stats['total_missing'],      delta_color="inverse")
    m6.metric("Overall Match %",           f"{stats['overall_match_pct']:.1f}%")

    st.markdown("##### 🔍 Missing Items")
    n1, n2, n3 = st.columns(3)
    n1.metric("Total Missing Items",          stats['total_missing_items'])
    n2.metric(f"Not Found in {target_name}",  stats['missing_in_target'],  delta_color="inverse")
    n3.metric(f"Not Found in {source_name}",  stats['missing_in_source'],  delta_color="inverse")

    st.markdown("##### 📋 Column-Level Statistics")
    c1, c2, c3 = st.columns(3)
    c1.metric("Total Columns Compared",  stats['total_columns_compared'])
    c2.metric("Total Matched Columns",   stats['total_matched_columns'])
    c3.metric("Total Unmatched Columns",
              stats['total_columns_compared'] - stats['total_matched_columns'])

    st.markdown("---")
    st.markdown("##### 📉 Visual Breakdown")
    chart1, chart2 = st.columns(2)

    with chart1:
        st.markdown("###### Items: Match Status Distribution")
        fig_pie = go.Figure(data=[go.Pie(
            labels=['100% Match', 'Has Issues', 'Missing Items'],
            values=[stats['total_100pct'],
                    stats['total_with_issues'],
                    stats['total_missing_items']],
            marker=dict(colors=['#90EE90', '#FFB6C1', '#FFE4B5']),
            hole=0.4,
            textinfo='label+percent+value',
            hovertemplate='%{label}: %{value} items (%{percent})<extra></extra>'
        )])
        fig_pie.update_layout(showlegend=True, height=400,
                              margin=dict(t=10, b=10, l=10, r=10))
        st.plotly_chart(fig_pie, use_container_width=True)

    with chart2:
        st.markdown("###### Issue Category Breakdown")
        if stats['issue_category_counts']:
            cat_df = pd.DataFrame(
                list(stats['issue_category_counts'].items()),
                columns=['Issue Category', 'Count']
            ).sort_values('Count', ascending=True)

            fig_bar = px.bar(
                cat_df, x='Count', y='Issue Category', orientation='h',
                color='Count',
                color_continuous_scale=[[0, '#FFE4B5'], [0.5, '#FFB6C1'], [1, '#FF6B6B']],
                text='Count'
            )
            fig_bar.update_traces(textposition='outside')
            fig_bar.update_layout(
                height=400, margin=dict(t=10, b=10, l=10, r=10),
                coloraxis_showscale=False,
                xaxis_title='Number of Issues', yaxis_title=''
            )
            st.plotly_chart(fig_bar, use_container_width=True)
        else:
            st.success("🎉 No issues found!")

    st.markdown("---")
    st.markdown("###### Items by Form Label")
    if stats['form_label_stats']:
        form_rows = []
        for fl, counts in stats['form_label_stats'].items():
            form_rows.append({
                'Form Label': fl,
                '100% Match': counts['match_100'],
                'Has Issues': counts['with_issues'],
            })
        form_df      = pd.DataFrame(form_rows).sort_values('Has Issues', ascending=False)
        chart_height = max(500, len(form_df) * 45 + 120)

        fig_form = px.bar(
            form_df, x='Form Label', y=['100% Match', 'Has Issues'],
            barmode='stack',
            color_discrete_map={'100% Match': '#90EE90', 'Has Issues': '#FFB6C1'},
            text_auto=True
        )
        fig_form.update_layout(
            height=chart_height,
            margin=dict(t=30, b=120, l=10, r=10),
            xaxis_title='Form Label', yaxis_title='Number of Items',
            legend_title='Status', xaxis_tickangle=-35, font=dict(size=13),
            legend=dict(orientation='h', yanchor='bottom', y=1.02,
                        xanchor='right', x=1)
        )
        fig_form.update_traces(textfont_size=13, textposition='inside')
        st.plotly_chart(fig_form, use_container_width=True)

    st.markdown("---")
    st.markdown("##### 📋 Breakdown by Form Label")
    if stats['form_label_stats']:
        form_detail_rows = []
        for fl, counts in stats['form_label_stats'].items():
            total = counts['total']
            pct   = (counts['match_100'] / total * 100) if total > 0 else 0
            form_detail_rows.append({
                'Form Label':   fl,
                'Total Items':  total,
                '100% Match':   counts['match_100'],
                'Has Issues':   counts['with_issues'],
                'Form Match %': f"{pct:.1f}%"
            })
        st.dataframe(
            pd.DataFrame(form_detail_rows).sort_values('Has Issues', ascending=False),
            use_container_width=True, hide_index=True
        )


# ─────────────────────────────────────────────
# REPORT GENERATION
# ─────────────────────────────────────────────

def create_comprehensive_report(
    all_comparisons, missing_in_source, missing_in_target,
    source_name, target_name, source_df
):
    output      = BytesIO()
    YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    summary_data = []
    issues_data  = []

    for tab_label, data in all_comparisons.items():
        comp_df   = data['comparison_df']
        mc        = comp_df['Match'].value_counts()
        total     = len(comp_df)
        matches_n = mc.get('match', 0)
        match_pct = (matches_n / total * 100) if total > 0 else 0

        cl_status = 'N/A'
        if data.get('codelist_comparison'):
            cl = data['codelist_comparison']
            cl_status = (f"✓ {cl['matches']} matches" if cl['status'] == 'match'
                         else f"✗ {cl['mismatches']} issues")

        ucl_status = 'N/A'
        if data.get('unit_codelist_comparison'):
            ucl = data['unit_codelist_comparison']
            ucl_status = (f"✓ {ucl['matches']} matches" if ucl['status'] == 'match'
                          else f"✗ {ucl['mismatches']} issues")

        item_name  = data['item_name']
        form_name  = data.get('form_name', '')
        form_label = data.get('form_label', '')
        ig_label   = data.get('item_group_label', '')

        summary_data.append({
            'Item Name': item_name, 'Form Name': form_name, 'Form Label': form_label,
            'Status': 'Compared', 'Match Type': data['match_type'],
            'Total Columns': total, 'Matches': matches_n,
            'Codelist Status': cl_status, 'Unit Codelist Status': ucl_status,
            'Match %': f"{match_pct:.1f}%",
        })

        if match_pct < 100.0:
            for _, row in comp_df[comp_df['Match'] != 'match'].iterrows():
                col_name     = row['Column Name']
                src_val      = row[f'{source_name} Value']
                tgt_val      = row[f'{target_name} Value']
                match_status = row['Match']
                issue_category, issue_column, issue_detail = parse_issue_fields(
                    match_status, col_name, src_val, tgt_val, source_name, target_name
                )
                issues_data.append({
                    'Item Name': item_name, 'Form Name': form_name,
                    'Form Label': form_label, 'Item Group Label': ig_label,
                    'Issue Category': issue_category,
                    'Issue Column': issue_column, 'Issue': issue_detail,
                })

    for item_name in missing_in_target:
        summary_data.append({
            'Item Name': item_name, 'Form Name': '', 'Form Label': '',
            'Status': f'Not found in {target_name}', 'Match Type': 'N/A',
            'Total Columns': 0, 'Matches': 0,
            'Codelist Status': 'N/A', 'Unit Codelist Status': 'N/A', 'Match %': 'N/A',
        })
        issues_data.append({
            'Item Name': item_name, 'Form Name': '', 'Form Label': '',
            'Item Group Label': '', 'Issue Category': f'Item Not Found in {target_name}',
            'Issue Column': '',
            'Issue': f"Item '{item_name}' exists in {source_name} but not in {target_name}",
        })

    for item_name in missing_in_source:
        summary_data.append({
            'Item Name': item_name, 'Form Name': '', 'Form Label': '',
            'Status': f'Not found in {source_name}', 'Match Type': 'N/A',
            'Total Columns': 0, 'Matches': 0,
            'Codelist Status': 'N/A', 'Unit Codelist Status': 'N/A', 'Match %': 'N/A',
        })
        issues_data.append({
            'Item Name': item_name, 'Form Name': '', 'Form Label': '',
            'Item Group Label': '', 'Issue Category': f'Item Not Found in {source_name}',
            'Issue Column': '',
            'Issue': f"Item '{item_name}' exists in {target_name} but not in {source_name}",
        })

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        summary_df = pd.DataFrame(summary_data)[[
            'Item Name', 'Form Name', 'Form Label', 'Status', 'Match Type',
            'Total Columns', 'Matches', 'Codelist Status', 'Unit Codelist Status', 'Match %',
        ]]
        summary_df.to_excel(writer, sheet_name='Comparison Summary', index=False)
        ws_summary = writer.book['Comparison Summary']
        _autofit_sheet(ws_summary)
        status_col_idx = summary_df.columns.get_loc('Status') + 1
        for row_idx in range(2, ws_summary.max_row + 1):
            ws_summary.cell(row=row_idx, column=status_col_idx).fill = YELLOW_FILL

        if issues_data:
            issues_df = pd.DataFrame(issues_data)[[
                'Item Name', 'Form Name', 'Form Label', 'Item Group Label',
                'Issue Category', 'Issue Column', 'Issue',
            ]]
            issues_df.to_excel(writer, sheet_name='Issues Only', index=False)
            _autofit_sheet(writer.book['Issues Only'])
        else:
            pd.DataFrame({'Message': ['All items have 100% match rate.']}).to_excel(
                writer, sheet_name='Issues Only', index=False)

    output.seek(0)
    return output


def _autofit_sheet(worksheet):
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=False, vertical='top', horizontal='left')
    for column in worksheet.columns:
        col_letter = column[0].column_letter
        max_len    = max((len(str(cell.value)) for cell in column if cell.value), default=0)
        worksheet.column_dimensions[col_letter].width = min(max_len + 2, 80)


# ─────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────

def _init_session_state():
    defaults = {
        'selected_items':          [],
        'ptd_df':                  None,
        'sds_df':                  None,
        'ptd_codelists_df':        None,
        'sds_codelists_df':        None,
        'ptd_unit_codelists_df':   None,
        'sds_unit_codelists_df':   None,
        'comparison_direction':    "PTD → SDS (Compare PTD columns against SDS)",
        'all_comparisons':         None,
        'comparison_complete':     False,
        'report_output':           None,
        'missing_in_source':       [],
        'missing_in_target':       [],
        'source_name':             '',
        'target_name':             '',
        'source_df':               None,
        # ── file identity tracking ─────────────────────────────────────
        # Store file name+size as a lightweight fingerprint.
        # _clear_comparison_state only fires when the fingerprint changes.
        'left_file_id':            None,
        'right_file_id':           None,
    }
    for key, val in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = val


def _file_id(uploaded_file):
    """Return a lightweight fingerprint for an uploaded file."""
    if uploaded_file is None:
        return None
    return (uploaded_file.name, uploaded_file.size)


def _clear_comparison_state():
    """Reset only comparison results — NOT file data or selections."""
    st.session_state.all_comparisons     = None
    st.session_state.comparison_complete = False
    st.session_state.report_output       = None


# ─────────────────────────────────────────────
# FILE LOADER HELPER
# ─────────────────────────────────────────────

def _load_file_data(uploaded_file, label, is_ptd):
    form_df = cl_df = ucl_df = None

    raw_form = parse_uploaded_file(
        uploaded_file, sheet_name='Form Definitions', is_ptd=is_ptd
    )
    if raw_form is not None:
        if is_ptd:
            result = process_ptd_dataframe(raw_form)
            if isinstance(result, tuple):
                form_df, orig, filt = result
                st.success(f"✅ {label} Form Definitions loaded: {len(form_df)} rows")
                if orig != filt:
                    st.info(f"ℹ️ Filtered from {orig} to {filt} rows (Y/Yes/Mod only)")
        else:
            form_df = process_sds_dataframe(raw_form)
            st.success(f"✅ {label} Form Definitions loaded: {len(form_df)} rows")

    raw_cl = parse_uploaded_file(uploaded_file, sheet_name='Codelists', is_ptd=False)
    if raw_cl is not None:
        result = process_choice_code_sheet(raw_cl, "Codelist")
        if result[0] is not None:
            cl_df, orig, filt = result
            st.success(f"✅ {label} Codelists loaded: {len(cl_df)} rows")
            if orig != filt:
                st.info(f"ℹ️ Codelists filtered from {orig} to {filt} rows")

    raw_ucl = parse_uploaded_file(
        uploaded_file, sheet_name='Unit Codelists', is_ptd=False
    )
    if raw_ucl is not None:
        result = process_choice_code_sheet(raw_ucl, "Unit Codelist")
        if result[0] is not None:
            ucl_df, orig, filt = result
            st.success(f"✅ {label} Unit Codelists loaded: {len(ucl_df)} rows")
            if orig != filt:
                st.info(f"ℹ️ Unit Codelists filtered from {orig} to {filt} rows")

    return form_df, cl_df, ucl_df


# ─────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────

def main():
    st.title("📊 PTD vs SDS Comparison Tool")
    _init_session_state()

    # ── Input method ──────────────────────────
    st.markdown("---")
    st.subheader("📥 Select Input Method")
    st.radio("How would you like to provide the data?",
             ["📁 Upload Excel Files"], horizontal=True)

    # ── Comparison direction ──────────────────
    st.markdown("---")
    direction_options = [
        "PTD → SDS (Compare PTD columns against SDS)",
        "SDS → PTD (Compare SDS columns against PTD)",
        "Parental SDS → Child SDS (Compare Parental SDS columns against Child SDS)",
        "Parental PTD → Child PTD (Compare Parental PTD columns against Child PTD)",
    ]
    comparison_direction = st.radio(
        "🔄 Select Comparison Direction:",
        direction_options,
        index=direction_options.index(st.session_state.comparison_direction)
    )
    # Only clear results if direction actually changed
    if comparison_direction != st.session_state.comparison_direction:
        st.session_state.comparison_direction = comparison_direction
        _clear_comparison_state()
    else:
        st.session_state.comparison_direction = comparison_direction

    if "Parental SDS → Child SDS" in comparison_direction:
        left_label, right_label   = "Parental SDS", "Child SDS"
        left_key,   right_key     = "parental_sds", "child_sds"
        is_left_ptd, is_right_ptd = False, False
    elif "Parental PTD → Child PTD" in comparison_direction:
        left_label, right_label   = "Parental PTD", "Child PTD"
        left_key,   right_key     = "parental_ptd", "child_ptd"
        is_left_ptd, is_right_ptd = True, True
    else:
        left_label, right_label = "PTD", "SDS"
        left_key,   right_key   = "ptd", "sds"
        is_left_ptd  = (left_label == "PTD")
        is_right_ptd = (right_label == "PTD")

    st.markdown("---")
    st.markdown("#### 📁 Upload Excel Files:")
    st.info("ℹ️ Data will be read from '**Form Definitions**', "
            "'**Codelists**', and '**Unit Codelists**' sheets.")

    if is_left_ptd and is_right_ptd:
        st.warning("⚠️ **For both PTD files:** First row skipped, "
                   "second row as headers. Data filtered for Y/Yes/Mod only.")
    elif is_left_ptd:
        st.warning(f"⚠️ **For {left_label} file:** First row skipped, "
                   "second row as headers. Data filtered for Y/Yes/Mod only.")
    elif is_right_ptd:
        st.warning(f"⚠️ **For {right_label} file:** First row skipped, "
                   "second row as headers. Data filtered for Y/Yes/Mod only.")

    st.info(
        "ℹ️ **Automatic Processing:**\n"
        "- PTD data filtered for Y/Yes/Mod in 'Used in trial' column\n"
        "- 'Modification comments' and 'Used in trial' columns removed\n"
        "- 'Decimal' column formatted (e.g., 1 → 1.0)\n"
        "- Codelists and Unit Codelists filtered to remove blank Choice Code rows\n"
        "- Codelist and Unit Codelist comparison for Name, Choice Code, and Choice Label"
    )
    st.markdown("---")

    # ── File uploaders ────────────────────────
    col1, col2 = st.columns(2)

    with col1:
        st.subheader(f"📄 {left_label} File")
        left_file = st.file_uploader(
            f"Upload {left_label} Excel file:",
            type=['xlsx', 'xls'], key=f"{left_key}_upload"
        )
        new_left_id = _file_id(left_file)
        # Only reload + clear if a DIFFERENT file was uploaded
        if left_file is not None and new_left_id != st.session_state.left_file_id:
            st.session_state.left_file_id = new_left_id
            _clear_comparison_state()
            with st.spinner(f"Reading {left_label} file..."):
                form_df, cl_df, ucl_df = _load_file_data(
                    left_file, left_label, is_left_ptd
                )
            st.session_state.ptd_df                = form_df
            st.session_state.ptd_codelists_df      = cl_df
            st.session_state.ptd_unit_codelists_df = ucl_df
        elif st.session_state.ptd_df is not None:
            st.success(f"✅ {left_label} data loaded (from session)")

    with col2:
        st.subheader(f"📄 {right_label} File")
        right_file = st.file_uploader(
            f"Upload {right_label} Excel file:",
            type=['xlsx', 'xls'], key=f"{right_key}_upload"
        )
        new_right_id = _file_id(right_file)
        # Only reload + clear if a DIFFERENT file was uploaded
        if right_file is not None and new_right_id != st.session_state.right_file_id:
            st.session_state.right_file_id = new_right_id
            _clear_comparison_state()
            with st.spinner(f"Reading {right_label} file..."):
                form_df, cl_df, ucl_df = _load_file_data(
                    right_file, right_label, is_right_ptd
                )
            st.session_state.sds_df                = form_df
            st.session_state.sds_codelists_df      = cl_df
            st.session_state.sds_unit_codelists_df = ucl_df
        elif st.session_state.sds_df is not None:
            st.success(f"✅ {right_label} data loaded (from session)")

    ptd_df                = st.session_state.ptd_df
    sds_df                = st.session_state.sds_df
    ptd_codelists_df      = st.session_state.ptd_codelists_df
    sds_codelists_df      = st.session_state.sds_codelists_df
    ptd_unit_codelists_df = st.session_state.ptd_unit_codelists_df
    sds_unit_codelists_df = st.session_state.sds_unit_codelists_df

    has_left  = ptd_df is not None
    has_right = sds_df is not None

    if not has_left and not has_right:
        st.warning(f"⚠️ Please provide both {left_label} and {right_label} data.")
        return
    if not has_left:
        st.warning(f"⚠️ Please provide {left_label} data.")
        return
    if not has_right:
        st.warning(f"⚠️ Please provide {right_label} data.")
        return

    st.markdown("---")
    st.success("✅ Both datasets loaded! Ready to compare.")

    c1, c2 = st.columns(2)
    with c1:
        if ptd_codelists_df is not None and sds_codelists_df is not None:
            st.success("✅ Codelists loaded for both files!")
        elif ptd_codelists_df is None and sds_codelists_df is None:
            st.warning("⚠️ Codelists not loaded. Codelist comparison will be skipped.")
        else:
            st.warning("⚠️ Codelists loaded for only one file.")
    with c2:
        if ptd_unit_codelists_df is not None and sds_unit_codelists_df is not None:
            st.success("✅ Unit Codelists loaded for both files!")
        elif ptd_unit_codelists_df is None and sds_unit_codelists_df is None:
            st.warning("⚠️ Unit Codelists not loaded. Unit Codelist comparison will be skipped.")
        else:
            st.warning("⚠️ Unit Codelists loaded for only one file.")

    if "PTD → SDS" in comparison_direction:
        source_df, target_df               = ptd_df, sds_df
        source_name, target_name           = "PTD", "SDS"
        source_codelists, target_codelists = ptd_codelists_df, sds_codelists_df
        source_ucl, target_ucl             = ptd_unit_codelists_df, sds_unit_codelists_df
    elif "SDS → PTD" in comparison_direction:
        source_df, target_df               = sds_df, ptd_df
        source_name, target_name           = "SDS", "PTD"
        source_codelists, target_codelists = sds_codelists_df, ptd_codelists_df
        source_ucl, target_ucl             = sds_unit_codelists_df, ptd_unit_codelists_df
    elif "Parental SDS → Child SDS" in comparison_direction:
        source_df, target_df               = ptd_df, sds_df
        source_name, target_name           = "Parental SDS", "Child SDS"
        source_codelists, target_codelists = ptd_codelists_df, sds_codelists_df
        source_ucl, target_ucl             = ptd_unit_codelists_df, sds_unit_codelists_df
    else:
        source_df, target_df               = ptd_df, sds_df
        source_name, target_name           = "Parental PTD", "Child PTD"
        source_codelists, target_codelists = ptd_codelists_df, sds_codelists_df
        source_ucl, target_ucl             = ptd_unit_codelists_df, sds_unit_codelists_df

    st.markdown("---")
    IGNORE_COLUMNS       = {'Definition Last Modified', 'Relationship Last Modified'}
    source_cols_filtered = [c for c in source_df.columns if c not in IGNORE_COLUMNS]

    m1, m2, m3, m4 = st.columns(4)
    m1.info(f"📊 {source_name} Records: {len(source_df)}")
    m2.info(f"📊 {target_name} Records: {len(target_df)}")
    m3.info(f"📋 {source_name} Columns: {len(source_cols_filtered)}")
    extra_lines = []
    if source_codelists is not None:
        extra_lines.append(f"📋 Codelists: {len(source_codelists)}")
    if source_ucl is not None:
        extra_lines.append(f"📋 Unit CL: {len(source_ucl)}")
    m4.info("\n".join(extra_lines) if extra_lines else "⚠️ No additional lists")

    st.markdown("---")

    source_items = get_unique_items(source_df)
    target_items = get_unique_items(target_df)

    only_in_source = sorted(source_items - target_items)
    only_in_target = sorted(target_items - source_items)

    if only_in_source or only_in_target:
        with st.expander(
            f"⚠️ Items Not Found in Both Files "
            f"({len(only_in_source) + len(only_in_target)} total)",
            expanded=False               # ← collapsed by default to avoid rerun issues
        ):
            st.info("ℹ️ These items are **excluded from the comparison tabs** below. "
                    "They will appear in the **Issues Only** sheet of the downloaded report.")
            ec1, ec2 = st.columns(2)
            with ec1:
                if only_in_source:
                    st.warning(f"**Only in {source_name} — not found in {target_name} "
                               f"({len(only_in_source)}):**")
                    st.dataframe(
                        pd.DataFrame({'Item Name': only_in_source,
                                      'Status': [f'Not found in {target_name}'] * len(only_in_source)}),
                        use_container_width=True, hide_index=True
                    )
            with ec2:
                if only_in_target:
                    st.warning(f"**Only in {target_name} — not found in {source_name} "
                               f"({len(only_in_target)}):**")
                    st.dataframe(
                        pd.DataFrame({'Item Name': only_in_target,
                                      'Status': [f'Not found in {source_name}'] * len(only_in_target)}),
                        use_container_width=True, hide_index=True
                    )

    st.markdown("---")

    items_in_both_names = source_items.intersection(target_items)
    source_item_keys    = [
        item for item in get_all_item_keys(source_df)
        if item['item_name'] in items_in_both_names
    ]
    option_map     = {item['tab_label']: item for item in source_item_keys}
    all_tab_labels = sorted(option_map.keys())

    st.subheader("Select Items to Compare")
    st.caption(
        f"ℹ️ Only items present in both files are shown here "
        f"({len(source_item_keys)} entries). "
        "Duplicate Item Names shown as 'Item Name (Form Label)'."
    )

    # ── Selection buttons — NO st.rerun() ────────────────────────────
    # Instead of rerunning, we set a flag and read it immediately below.
    sc1, sc2, sc3 = st.columns(3)
    with sc1:
        if st.button("✅ Select All", use_container_width=True):
            st.session_state.selected_items = all_tab_labels
    with sc2:
        if st.button("❌ Clear Selection", use_container_width=True):
            st.session_state.selected_items = []
    with sc3:
        if st.button(f"🔍 Select from {source_name} only", use_container_width=True):
            st.session_state.selected_items = all_tab_labels

    # Validate stored selection against current option list
    valid_default = [l for l in st.session_state.selected_items if l in all_tab_labels]

    selected_labels = st.multiselect(
        "Select Item Names:",
        options=all_tab_labels,
        default=valid_default,
        key='multiselect_items'
    )
    # Sync back without triggering a rerun
    st.session_state.selected_items = selected_labels

    if selected_labels:
        st.info(f"📋 Selected {len(selected_labels)} item(s)")
    else:
        st.warning("⚠️ No items selected")

    # ── Compare button ────────────────────────────────────────────────
    if selected_labels and st.button(
        "🔍 Compare Selected Items", type="primary", use_container_width=True
    ):
        all_comparisons = {}

        with st.spinner("Building lookup indices..."):
            src_simple, src_composite, src_dups = build_lookup_dictionaries(source_df)
            tgt_simple, tgt_composite, tgt_dups = build_lookup_dictionaries(target_df)
            source_columns = [c for c in source_df.columns if c not in IGNORE_COLUMNS]

        codelist_col      = find_column(source_df, 'codelist')
        unit_codelist_col = find_column(source_df, 'unit codelist')

        progress_bar = st.progress(0)
        status_text  = st.empty()
        total_items  = len(selected_labels)
        start_time   = time.time()

        for idx, tab_label in enumerate(selected_labels):
            item_info = option_map[tab_label]
            if idx % 20 == 0 or idx == total_items - 1:
                status_text.text(f"Comparing {idx + 1}/{total_items}: {tab_label}")
                progress_bar.progress((idx + 1) / total_items)

            item_name        = item_info['item_name']
            form_label       = item_info['form_label']
            item_group_label = item_info['item_group_label']

            source_row, target_row, match_type = find_matching_rows_optimized(
                item_name,
                src_simple, src_composite, src_dups,
                tgt_simple, tgt_composite, tgt_dups,
                form_label, item_group_label
            )

            if source_row is None and target_row is None:
                continue

            form_name_val  = source_row.get('Form Name', '')        if source_row is not None else ''
            form_label_val = source_row.get('Form Label', '')       if source_row is not None else ''
            ig_label_val   = source_row.get('Item Group Label', '') if source_row is not None else ''

            codelist_comparison = None
            if codelist_col and source_row is not None:
                cl_val = source_row.get(codelist_col)
                if pd.notna(cl_val) and str(cl_val).strip():
                    codelist_comparison = compare_codelists(
                        cl_val, source_codelists, target_codelists,
                        source_name, target_name, "Codelist"
                    )

            unit_codelist_comparison = None
            if unit_codelist_col and source_row is not None:
                ucl_val = source_row.get(unit_codelist_col)
                if pd.notna(ucl_val) and str(ucl_val).strip():
                    unit_codelist_comparison = compare_codelists(
                        ucl_val, source_ucl, target_ucl,
                        source_name, target_name, "Unit Codelist"
                    )

            comparison_df = create_comparison_dataframe(
                source_row, target_row, source_columns,
                source_name, target_name,
                codelist_comparison, unit_codelist_comparison
            )

            all_comparisons[tab_label] = {
                'comparison_df':            comparison_df,
                'match_type':               match_type,
                'source_exists':            source_row is not None,
                'target_exists':            target_row is not None,
                'codelist_comparison':      codelist_comparison,
                'unit_codelist_comparison': unit_codelist_comparison,
                'item_name':                item_name,
                'form_name':                form_name_val,
                'form_label':               form_label_val,
                'item_group_label':         ig_label_val,
            }

        elapsed = time.time() - start_time
        status_text.text(f"✅ Comparison complete! ({elapsed:.2f}s)")
        progress_bar.progress(1.0)

        # ── Persist everything in session state ───────────────────────
        st.session_state.all_comparisons     = all_comparisons
        st.session_state.comparison_complete = True
        st.session_state.source_name         = source_name
        st.session_state.target_name         = target_name
        st.session_state.source_df           = source_df
        st.session_state.missing_in_source   = only_in_target
        st.session_state.missing_in_target   = only_in_source
        st.session_state.report_output       = None

        st.success(
            f"✅ Completed comparison for {len(all_comparisons)} items in {elapsed:.2f}s!"
        )

    # ── Results — always rendered from session state ──────────────────
    # This block runs on EVERY rerun as long as results exist in state.
    if st.session_state.comparison_complete and st.session_state.all_comparisons:

        all_comparisons = st.session_state.all_comparisons
        s_name          = st.session_state.source_name
        t_name          = st.session_state.target_name

        st.markdown("---")
        st.subheader("📊 Comparison Summary")

        summary_rows = []
        for tab_label, data in all_comparisons.items():
            comp_df   = data['comparison_df']
            mc        = comp_df['Match'].value_counts()
            total     = len(comp_df)
            matches_n = mc.get('match', 0)
            match_pct = (matches_n / total * 100) if total > 0 else 0

            cl_info = "N/A"
            if data.get('codelist_comparison'):
                cl = data['codelist_comparison']
                cl_info = (f"✓ {cl['matches']} matches" if cl['status'] == 'match'
                           else (cl['message'] if cl['status'] in
                                 ('missing_source', 'missing_target', 'not_found')
                                 else f"✗ {cl['mismatches']} issues"))

            ucl_info = "N/A"
            if data.get('unit_codelist_comparison'):
                ucl = data['unit_codelist_comparison']
                ucl_info = (f"✓ {ucl['matches']} matches" if ucl['status'] == 'match'
                            else (ucl['message'] if ucl['status'] in
                                  ('missing_source', 'missing_target', 'not_found')
                                  else f"✗ {ucl['mismatches']} issues"))

            summary_rows.append({
                'Item Name':               data['item_name'],
                'Form Label':              data['form_label'],
                'Match Type':              data['match_type'],
                'Total Columns':           total,
                'Matches ✅':              matches_n,
                'Mismatches ❌':           mc.get('mismatch', 0),
                f'Missing in {t_name} ⚠️': mc.get('missing_target', 0),
                f'Missing in {s_name} ⚠️': mc.get('missing_source', 0),
                'Codelist':                cl_info,
                'Unit Codelist':           ucl_info,
                'Match %':                 f"{match_pct:.1f}%",
            })

        st.dataframe(pd.DataFrame(summary_rows),
                     use_container_width=True, height=300)

        # ── Statistics ─────────────────────────────────────────────────
        stats = compute_statistics(
            all_comparisons,
            st.session_state.missing_in_source,
            st.session_state.missing_in_target,
            s_name, t_name
        )
        render_statistics(stats, s_name, t_name)

        # ── Detailed Results ───────────────────────────────────────────
        items_100pct      = []
        items_with_issues = {}

        for tab_label, data in all_comparisons.items():
            comp_df   = data['comparison_df']
            matches_n = (comp_df['Match'] == 'match').sum()
            pct       = (matches_n / len(comp_df) * 100) if len(comp_df) > 0 else 0
            if pct == 100.0:
                items_100pct.append(tab_label)
            else:
                items_with_issues[tab_label] = data

        st.markdown("---")
        st.subheader("📋 Detailed Results")

        if items_100pct:
            st.success(f"✅ {len(items_100pct)} item(s) with 100% match (hidden)")
            with st.expander(f"View {len(items_100pct)} items with 100% match"):
                st.write(items_100pct)

        if not items_with_issues:
            st.info("🎉 All items have 100% match!")
        else:
            st.info(f"Showing all {len(items_with_issues)} item(s) with discrepancies")
            tabs = st.tabs(list(items_with_issues.keys()))

            for tab, (tab_label, data) in zip(tabs, items_with_issues.items()):
                with tab:
                    comparison_df = data['comparison_df']
                    mc        = comparison_df['Match'].value_counts()
                    matches_n = mc.get('match', 0)
                    match_rate = (matches_n / len(comparison_df) * 100
                                  if len(comparison_df) > 0 else 0)

                    tc1, tc2 = st.columns(2)
                    with tc1:
                        if data['match_type']:
                            st.success(f"✅ Matched by: **{data['match_type']}**")
                        st.info(f"In {s_name}: {'✓' if data['source_exists'] else '✗'}")
                        if data['form_label']:
                            st.info(f"Form Label: **{data['form_label']}**")
                        if data.get('codelist_comparison'):
                            cl = data['codelist_comparison']
                            (st.success if cl['status'] == 'match' else st.warning)(
                                f"{'✓' if cl['status'] == 'match' else '⚠️'} "
                                f"Codelist: {cl['message']}"
                            )
                        if data.get('unit_codelist_comparison'):
                            ucl = data['unit_codelist_comparison']
                            (st.success if ucl['status'] == 'match' else st.warning)(
                                f"{'✓' if ucl['status'] == 'match' else '⚠️'} "
                                f"Unit Codelist: {ucl['message']}"
                            )
                    with tc2:
                        st.metric("Match Rate", f"{match_rate:.1f}%")
                        st.info(f"In {t_name}: {'✓' if data['target_exists'] else '✗'}")

                    rc1, rc2, rc3, rc4 = st.columns(4)
                    rc1.metric("Total Rows",    len(comparison_df))
                    rc2.metric("✅ Matches",    mc.get('match', 0))
                    rc3.metric("❌ Mismatches", mc.get('mismatch', 0))
                    rc4.metric("⚠️ Missing",
                               mc.get('missing_source', 0) + mc.get('missing_target', 0))

                    st.markdown("---")
                    st.markdown("##### Complete Comparison Table")
                    st.dataframe(
                        comparison_df.style.apply(highlight_differences, axis=1),
                        use_container_width=True, height=500
                    )

        st.markdown("---")
        st.markdown("**Legend:**")
        lc1, lc2, lc3, lc4 = st.columns(4)
        lc1.markdown("🟢 **Green**: Values match")
        lc2.markdown("🔴 **Red**: Values differ")
        lc3.markdown("🟡 **Orange**: Value missing")
        lc4.markdown("📋 **Codelist/Unit CL rows**: Choice Code comparisons")

        # ── Download ───────────────────────────────────────────────────
        st.markdown("---")
        st.subheader("📥 Download Report")
        dc1, dc2 = st.columns([2, 1])
        with dc1:
            st.info("📊 Report includes:\n"
                    "- **Sheet 1**: Comparison Summary\n"
                    "- **Sheet 2**: Issues Only")
        with dc2:
            if st.session_state.report_output is None:
                with st.spinner("Generating report..."):
                    st.session_state.report_output = create_comprehensive_report(
                        all_comparisons,
                        st.session_state.missing_in_source,
                        st.session_state.missing_in_target,
                        s_name, t_name,
                        st.session_state.source_df
                    )
            st.download_button(
                label="📥 Download Report",
                data=st.session_state.report_output,
                file_name=f"comparison_{s_name}_vs_{t_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )


if __name__ == "__main__":
    main()
